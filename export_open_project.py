#!/usr/bin/env python3
import argparse
import calendar
import json
import os
import re
from datetime import date
from urllib.parse import urljoin, urlparse

import requests
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ISO_DURATION_RE = re.compile(r"^P(?:\d+Y)?(?:\d+M)?(?:\d+D)?"
                             r"(?:T(?:(?P<h>\d+)H)?(?:(?P<m>\d+)M)?(?:(?P<s>\d+)S)?)?$")

def iso_duration_to_hours(s: str) -> float:
    """Convert ISO-8601 duration like 'PT5H30M' to decimal hours."""
    if not s:
        return 0.0
    m = ISO_DURATION_RE.match(s)
    if not m:
        return 0.0
    h = int(m.group('h') or 0)
    mnts = int(m.group('m') or 0)
    sec = int(m.group('s') or 0)
    return round(h + mnts / 60 + sec / 3600, 2)

def month_bounds(ym: str):
    """Return first_date, last_date for 'YYYY-MM'."""
    year, month = map(int, ym.split("-"))
    first = date(year, month, 1)
    last = date(year, month, calendar.monthrange(year, month)[1])
    return first, last

def get_collection(session: requests.Session, url: str):
    """Iterate through a HAL paginated collection using nextByOffset links."""
    while url:
        r = session.get(url)
        r.raise_for_status()
        data = r.json()
        for el in data.get("_embedded", {}).get("elements", []):
            return_url = url  # keep for errors
            yield el
        next_link = (data.get("_links", {}).get("nextByOffset") or {}).get("href")
        url = urljoin(session.base_url, next_link) if next_link else None

def parse_entity_id(href: str) -> str:
    """Extract numeric ID from /api/v3/<resource>/<id>."""
    try:
        path = urlparse(href).path.strip("/")
        return path.split("/")[-1]
    except Exception:
        return ""

def resolve_custom_option_value(session: requests.Session, href: str) -> str:
    """Given a /api/v3/custom_options/{id} href, return the option's value (cached)."""
    if not href:
        return ""
    if not hasattr(session, "_custom_opt_cache"):
        session._custom_opt_cache = {}
    if href in session._custom_opt_cache:
        return session._custom_opt_cache[href]
    url = urljoin(session.base_url, href)
    r = session.get(url)
    r.raise_for_status()
    val = r.json().get("value", "")
    session._custom_opt_cache[href] = val
    return val

def main():
    p = argparse.ArgumentParser(description="Export OpenProject monthly timesheet to Excel")
    p.add_argument("--base-url", default=os.getenv("OPENPROJECT_BASE_URL"), required=False,
                   help="e.g. https://openproject.example.com")
    p.add_argument("--api-key", default=os.getenv("OPENPROJECT_API_KEY"), required=False,
                   help="Your OpenProject API key")
    p.add_argument("--month", "-m", default=date.today().strftime("%Y-%m"),
                   help="Month to export, format YYYY-MM (default: current month)")
    p.add_argument("--user", default=os.getenv("OPENPROJECT_USER", "me"),
                   help="User id or 'me' (default: me)")
    p.add_argument("--location-cf", default=os.getenv("OPENPROJECT_LOCATION_CF", ""),
                   help="Custom field key for Location, e.g. customField7 (optional)")
    p.add_argument("--out", default=None, help="Output XLSX path (default: ./timesheet-YYYY-MM.xlsx)")
    p.add_argument("--page-size", type=int, default=200, help="API page size (default: 200)")
    args = p.parse_args()

    if not args.base_url or not args.api_key:
        raise SystemExit("Set --base-url and --api-key (or OPENPROJECT_BASE_URL / OPENPROJECT_API_KEY).")

    first, last = month_bounds(args.month)
    out_path = args.out or f"timesheet-{args.month}.xlsx"

    session = requests.Session()
    session.auth = ("apikey", args.api_key)  # per API docs
    # stash base_url on session for pagination helpers
    session.base_url = args.base_url.rstrip("/")

    filters = [
        {"spent_on": {"operator": "<>d", "values": [first.isoformat(), last.isoformat()]}},
        {"user_id": {"operator": "=", "values": [str(args.user)]}},
    ]
    params = {
        "filters": json.dumps(filters),
        "pageSize": args.page_size,
        "sortBy": json.dumps([["spent_on", "asc"], ["id", "asc"]]),
    }
    list_url = f"{session.base_url}/api/v3/time_entries"
    query_url = requests.Request("GET", list_url, params=params).prepare().url

    rows = []
    for te in get_collection(session, query_url):
        spent_on = te.get("spentOn")
        hours = iso_duration_to_hours(te.get("hours"))
        comment = (te.get("comment") or {}).get("raw") or ""
        comment = " ".join(comment.split())  # squash newlines/extra spaces

        links = te.get("_links", {})
        entity = links.get("entity") or {}
        activity = links.get("activity") or {}

        entity_href = entity.get("href", "")
        assignment_num = parse_entity_id(entity_href)  # fall back to empty if missing
        activity_name = activity.get("title") or ""   # usually available inline

        # Location custom field handling:
        location = "remote"  # default to remote
        cf_key = args.location_cf.strip()
        if cf_key:
            # try direct property (text/date/integer types)
            if cf_key in te:
                location = te.get(cf_key) or "remote"
            else:
                # try as link (list-type custom field)
                cf_link = (links.get(cf_key) or {}).get("href", "")
                if cf_link:
                    location = resolve_custom_option_value(session, cf_link) or "remote"

        composed = f"{assignment_num}_{activity_name}_{comment}".strip("_")
        rows.append({
            "Date": spent_on,
            "working hours": hours,
            "Location": location,
            "Assignment number_Activity_Work content": composed,
        })

    # Build DataFrame and write Excel
    df = pd.DataFrame(rows, columns=[
        "Date", "working hours", "Location", "Assignment number_Activity_Work content"
    ])
    # If you want one row per day with summed hours, uncomment below:
    # df = (df.groupby(["Date", "Location"], as_index=False)
    #         .agg({"working hours": "sum"}))

    # Excel export with formatting
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet_name = args.month
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Get the worksheet to apply formatting
        ws = writer.sheets[sheet_name]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Format header row
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # Format data rows
        for row_num in range(2, len(df) + 2):
            # Date column - center aligned
            cell = ws.cell(row=row_num, column=1)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

            # Working hours column - center aligned
            cell = ws.cell(row=row_num, column=2)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

            # Location column - center aligned
            cell = ws.cell(row=row_num, column=3)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

            # Assignment content column - left aligned with text wrap
            cell = ws.cell(row=row_num, column=4)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border_style

        # Set column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 14  # Working hours
        ws.column_dimensions['C'].width = 10  # Location
        ws.column_dimensions['D'].width = 80  # Assignment content

        # Set header row height
        ws.row_dimensions[1].height = 30

    print(f"Wrote {len(df)} rows to {out_path}")

if __name__ == "__main__":
    main()

